import tkinter as tk
from tkinter import messagebox  # import this to fix messagebox error
import pymongo #用來操作mongodb資料庫系統或與它溝通的模組(或稱套件)
from openpyxl import load_workbook,Workbook
from tkinter import filedialog
import pathlib
# import os

myclient = pymongo.MongoClient("mongodb://localhost:27017/") #連結mongodb #pymongo.MongoClient(host='localhost', port=27017)
mydb = myclient["db_manager"] #在mongodb中建立一個名為test2的資料庫
mycol = mydb["users"] #在test2中建一個名為users的集合(資料表)，在mongodb系統中，把資料表叫做集合(collection)，在許多其它資料庫系統中稱為資料表(table)

window = tk.Tk()
window.title('Welcome')
window.geometry('500x600') #('350x200')
rootFrame=tk.Frame(window)
rootFrame.pack(pady=20)

user_name = tk.StringVar() #產生一個StringVar 的物件
user_account = tk.StringVar()
user_pwd = tk.StringVar()
inquireVar = tk.StringVar()


def saveDB():
    if(user_name.get()=='' or user_account.get()=='' or user_pwd.get()==''): #兩條件式只要有一個為真，or起來就為真
        tk.messagebox.showerror(message='所有資料不能為空')
    else:
        curselections=listbox.curselection() #得到所有反白資料的索引值(串列)
        data = {'name':user_name.get(), 'account':user_account.get(), 'password':user_pwd.get()}
        index=0
        if(len(curselections)>0): #有選擇listbox上的項目時所做的操作
            str1=listbox.get('anchor')
            arr=curselections
            index=arr[0]
            if(str1==listbox.get(arr[0])): index=arr[-1]
            elif(str1==listbox.get(arr[-1])): index=arr[0]
            listbox.selection_clear(first=arr[0],last=arr[-1]) #刪掉所有反白(無論反白一行或多行，修改後都把反白刪掉)
            
        for i in range(index,listbox.size()):
            if(eval(listbox.get(i))['account']==data['account']):
                listbox.delete(first=i) 
                listbox.insert(i, data) #, str(data))
                break

        if(mycol.find_one({'account':data['account']})!=None):
            myquery = { "account": data['account'] } #這是update_many的第一個參數，也就是要找到資料的檢索條件
            newvalues = { "$set": data } #這是update_many的第二個參數，也就是要更新進去的資料
            mycol.update_one(myquery, newvalues)
        else:
            listbox.insert('end', data)
            mycol.insert_one(data)


def readDB():
    listbox.delete(first=0, last=listbox.size()-1)
    for x in mycol.find({},{"_id": 0}): #不返回"_id"資料。讀出users資料表(也就是mycol)中的所有資料，find()就是查詢資料庫的函數，若要查詢特定的資料，那麼就要給予檢索條件，比如find({account:'abc'})，若沒有給條件，代表要查詢出所有資料
        listbox.insert('end', x)
    if(listbox.size()==0): listbox.insert('end', '目前資料庫沒有資料')


def deleteData():
    arr=listbox.curselection() #得到反白選取到之資料的索引值
    if(len(arr)==0): #都沒有選，詢問是否刪除全部資料
        res = tk.messagebox.askyesno('刪除資料','請問是否要刪除全部資料，若否，請選擇資料後再按刪除')
        if(res): #
            listbox.delete(first=0,last='end') #刪listbox所有資料
            mycol.delete_many({}) #刪資料庫所有資料
            return

    for i in range(len(arr)-1,-1,-1): #倒著刪才不用管序號問題
        data=listbox.get(arr[i])
        data = eval(data) #轉dict
        query = {'account': data['account']} #刪資料庫中的一筆資料
        mycol.delete_one(query)
        listbox.delete(first=arr[i]) #刪listbox上的一筆資料
    #另一種寫法：
    # for i in range(listbox.size()-1,-1,-1):
    #     if(listbox.selection_includes(i)): #第i行是否已選
    #         query = {'account': eval(listbox.get(i))['account']} #刪資料庫中的一筆資料
    #         mycol.delete_one(query)
    #         listbox.delete(first=i) #刪listbox上的一筆資料


def uploadData():
    file = filedialog.askopenfilename(filetypes = [("檔案","*.xlsx")]) #叫出檔案總管視窗，可filetypes去限定所要顯現的檔案類型(副檔名)
    if(file==''):
        tk.messagebox.showinfo(title='讀檔結果', message = '你沒有選擇檔案')
        return #在函式中遇到return 就會跳出函式(不會再往下執行)

    wb = load_workbook(file) #載入檔案，比如：'usersData.xlsx'
    sheet = wb.active #通過active屬性來訪問工作表(表單)

    listbox.delete(first=0, last='end') #先把listbox 上所有資料刪掉再重新添加
    mycol.delete_many({}) #刪資料庫所有資料
    for row in sheet.iter_rows(min_row=2, min_col=1): ##讀資料，從第2列第0行開始遍歷，遍歷所有行
        # for cell in row: print(cell, cell.value)
        # row=list(row)
        # print(row)
        data = {'name':row[0].value, 'account':row[1].value, 'password':row[2].value}
        listbox.insert('end', data)
        mycol.insert_one(data)


def saveAs():
    filePath = filedialog.asksaveasfilename(title=u'另存excel檔', filetypes = [("檔案","*.xlsx")])
    if(filePath==''):
        tk.messagebox.showinfo(title='讀檔結果', message = '你沒有選擇檔案')
    else:
        path = pathlib.Path(filePath)
        fileName=path.name #os.path.basename(filePath)
        if(path.suffix==''): #if(''.join(path.suffixes)==''): //if(os.path.splitext(filePath)[1]==''): #得到副檔名
            fileName+='.xlsx'
        # print('儲存檔案：',fileName)
        wb=Workbook() #創建工作表
        ws = wb.active #通過active屬性來訪問工作表(表單)：
        ws.title = "Sheet1" #利用title屬性設定表單名稱
        sheet = wb['Sheet1'] #得到名稱為"Sheet1"的表單
        #在表單寫入資料
        sheet.cell(row=1, column=1, value="name")
        sheet.cell(row=1, column=2, value="account")
        sheet.cell(row=1, column=3, value="password")
        for i in range(listbox.size()):
            data = eval(listbox.get(i))
            sheet.cell(row=i+2, column=1, value=data['name'])
            sheet.cell(row=i+2, column=2, value=data['account'])
            sheet.cell(row=i+2, column=3, value=data['password'])
        wb.save(fileName)


inquireFlag=False
def inquireData(code, index, substr):
    global inquireFlag
    inquireFlag=True
    return 1
def keyPressedHandler(e):
    global inquireFlag
    if inquireFlag:
        inquireFlag=False #判斷過後就回復False，才能重新使用
        data=inquireVar.get()
        # print(data)
        regexp={"$regex":data, "$options":"$i"} #在SQL中等同 eval("/"+data+"/i")
        query0 = { "name": regexp }
        query1 = { "account": regexp }
        query2 = { "password": regexp }
        totalQuery={"$or": [query0, query1, query2]}

        listbox.delete(first=0, last='end')
        for x in mycol.find(totalQuery,{"_id": 0}): #不返回"_id"資料。讀出users資料表(也就是mycol)中的所有資料，find()就是查詢資料庫的函數，若要查詢特定的資料，那麼就要給予檢索條件，比如find({account:'abc'})，若沒有給條件，代表要查詢出所有資料
            listbox.insert('end', x)
window.bind('<KeyPress>',keyPressedHandler) #事件綁定或處理函式


def select_list_item(e):
    w = e.widget
    str1=w.get('anchor') #錨點(無論是由上而下還是由下而上反白，都是得到反白資料的第一行)
    # print(str1)
    curselections=w.curselection() #得到所有選擇(反白)資料的索引值(串列)
    index=curselections[0]
    if(str1==w.get(curselections[0])): index=curselections[-1] #由上往下拉，索引值index就設為curselections的最後一個元素值
    elif(str1==w.get(curselections[-1])): index=curselections[0] #由下往上拉，索引值就設為curselections的第一個元素值
    # print(index,curselections)
    data = w.get(index) #字串型別
    # print(type(data),data,data[10:14])
    data = eval(data) #轉dict，為什麼要轉，因為用字串截取的方式太難了(data[ : ])
    # print(type(data),data,data['name'])
    user_name.set(data['name']) #在js 是用 data.name(請對照比較)
    user_account.set(data['account']) #在js 是用 data.account
    user_pwd.set(data['password']) #在js 是用 data.password


frame1=tk.Frame(rootFrame)
frame1.pack()
tk.Label(frame1, text='Name: ').grid(row=0,column=0)
entry_user_name = tk.Entry(frame1, textvariable=user_name) #把StringVar物件設定給textvariable屬性，如此就可以用get()函數來得到entry元件所輸入的字串，也可以用set()函數把字串設定進去
entry_user_name.grid(row=0, column=1)
tk.Label(frame1, text='Account: ').grid(row=1, column=0)
entry_user_account = tk.Entry(frame1, textvariable=user_account) #, validate='key', validatecommand=account_check)
entry_user_account.grid(row=1, column=1, pady=10)
tk.Label(frame1, text='Password: ').grid(row=2, column=0)
entry_user_pwd = tk.Entry(frame1, textvariable=user_pwd) #, show='*')
entry_user_pwd.grid(row=2, column=1)

frame2=tk.Frame(rootFrame)
frame2.pack(pady=20)
btn_save = tk.Button(frame2, text='儲存', command=saveDB).pack(side='left')
btn_read = tk.Button(frame2, text='讀取', command=readDB).pack(side='left', padx=15)
btn_delete = tk.Button(frame2, text='刪除', command=deleteData).pack(side='left')
btn_upload = tk.Button(frame2, text='上傳excel資料', command=uploadData).pack(side='left', padx=15)
tk.Button(frame2, text='另存excel檔', command=saveAs).pack(side='left')

frame3=tk.Frame(rootFrame)
frame3.pack()
tk.Label(frame3, text='查詢: ').pack(side='left')
check = window.register(inquireData)
entry_inquireData = tk.Entry(frame3, textvariable=inquireVar, validate='key', validatecommand=(check, '%d', '%i', '%S')).pack(side='left')

frame4=tk.Frame(rootFrame)
frame4.pack(pady=5)
scrollbar = tk.Scrollbar(frame4)
scrollbar.pack(side='right', fill='y') #side='right' 放入右邊。fill='y' 向 y 軸填滿
scrollbar2 = tk.Scrollbar(frame4,orient='horizontal')
scrollbar2.pack(side='bottom', fill='x')
listbox = tk.Listbox(frame4, selectmode='extended', xscrollcommand=scrollbar2.set, yscrollcommand=scrollbar.set, width=60)
listbox.pack(fill='both', expand=1) #fill='both'向x和y軸填滿。expand=1開啟fill #因listbox放在frame4中，所以fill與expand的效果出不來
scrollbar.config(command=listbox.yview) #scrollbar 移動時使 listbox 跟著移動
scrollbar2.config(command=listbox.xview) #scrollbar 移動時使 listbox 跟著移動

for x in mycol.find({},{"_id": 0}): #不返回"_id"資料。讀出users資料表(也就是mycol)中的所有資料，find()就是查詢資料庫的函數，若要查詢特定的資料，那麼就要給予檢索條件，比如find({account:'abc'})，若沒有給條件，代表要查詢出所有資料
    listbox.insert('end', x)
if(listbox.size()==0): listbox.insert('end', '目前資料庫沒有資料')
listbox.bind('<ButtonRelease>', select_list_item)

window.mainloop()